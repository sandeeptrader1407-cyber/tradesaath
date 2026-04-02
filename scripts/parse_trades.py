#!/usr/bin/env python3
"""
TradeSaath Local Trade Parser (Python)
Parses PDF, CSV, TSV, XLS/XLSX trade files WITHOUT AI
Extracts trades, pairs buy/sell, calculates all KPIs
Returns structured JSON

Usage:
    python parse_trades.py <file_path> [--output result.json]

Requirements:
    pip install pdfplumber openpyxl
"""

import sys
import os
import json
import csv
import re
import io
from datetime import datetime
from typing import Optional

# ─── Column name patterns ───
COL_PATTERNS = {
    'time': re.compile(r'^(time|trade.?time|exec.?time|order.?time|timestamp|executed.?at)', re.I),
    'symbol': re.compile(r'^(symbol|scrip|instrument|stock|name|contract|underlying|security)', re.I),
    'side': re.compile(r'^(side|type|trade.?type|buy.?sell|action|order.?type|b.?s|direction)', re.I),
    'qty': re.compile(r'^(qty|quantity|lots|volume|traded.?qty|net.?qty|filled)', re.I),
    'price': re.compile(r'^(price|rate|avg.?price|trade.?price|executed.?price|avg.?rate)', re.I),
    'amount': re.compile(r'^(amount|value|net.?amount|turnover|total|net.?total)', re.I),
    'buy_qty': re.compile(r'^(buy.?qty|buy.?quantity|buy.?vol)', re.I),
    'sell_qty': re.compile(r'^(sell.?qty|sell.?quantity|sell.?vol)', re.I),
    'buy_price': re.compile(r'^(buy.?price|buy.?rate|buy.?avg|buy.?value)', re.I),
    'sell_price': re.compile(r'^(sell.?price|sell.?rate|sell.?avg|sell.?value)', re.I),
    'pnl': re.compile(r'^(pnl|p.?&.?l|profit|loss|net.?pnl|realized|realised|net.?profit)', re.I),
    'date': re.compile(r'^(date|trade.?date|order.?date|exec.?date)', re.I),
    'expiry': re.compile(r'^(expiry|expiry.?date|exp)', re.I),
    'strike': re.compile(r'^(strike|strike.?price)', re.I),
    'opt_type': re.compile(r'^(option.?type|opt.?type|ce.?pe|call.?put)', re.I),
}


def detect_broker(text: str) -> str:
    t = text.lower()
    brokers = [
        ('zerodha', 'Zerodha'), ('kite', 'Zerodha'), ('fyers', 'Fyers'),
        ('angel', 'Angel One'), ('groww', 'Groww'), ('upstox', 'Upstox'),
        ('icici', 'ICICI Direct'), ('hdfc', 'HDFC Securities'),
        ('kotak', 'Kotak Securities'), ('5paisa', '5Paisa'), ('dhan', 'Dhan'),
        ('paytm', 'Paytm Money'), ('motilal', 'Motilal Oswal'),
        ('sharekhan', 'Sharekhan'), ('finvasia', 'Finvasia'),
        ('flattrade', 'Flattrade'), ('interactive', 'Interactive Brokers'),
        ('robinhood', 'Robinhood'), ('shoonya', 'Finvasia'),
    ]
    for key, name in brokers:
        if key in t:
            return name
    return 'Unknown'


def detect_market(text: str) -> str:
    t = text.lower()
    if any(k in t for k in ['nse', 'bse', 'nifty', 'banknifty', 'sensex']):
        return 'NSE'
    if any(k in t for k in ['nyse', 'nasdaq']):
        return 'NYSE'
    if 'forex' in t or 'fx' in t:
        return 'Forex'
    return 'NSE'


def detect_currency(text: str) -> str:
    t = text.lower()
    if any(k in t for k in ['inr', '₹', 'rs', 'rupee']):
        return 'INR'
    if any(k in t for k in ['usd', '$']):
        return 'USD'
    return 'INR'


def detect_date(text: str) -> str:
    patterns = [
        r'(\d{4}[-/]\d{2}[-/]\d{2})',
        r'(\d{2}[-/]\d{2}[-/]\d{4})',
    ]
    for p in patterns:
        m = re.search(p, text)
        if m:
            d = m.group(1).replace('/', '-')
            if re.match(r'^\d{4}', d):
                return d
            parts = d.split('-')
            if len(parts) == 3 and len(parts[2]) == 4:
                return f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
            return d
    return datetime.now().strftime('%Y-%m-%d')


def classify_session(time_str: str) -> str:
    try:
        h = int(time_str.split(':')[0])
        if h < 11:
            return 'morning'
        elif h < 14:
            return 'midday'
        return 'afternoon'
    except (ValueError, IndexError):
        return 'morning'


def time_gap_minutes(t1: str, t2: str) -> Optional[float]:
    try:
        h1, m1 = map(int, t1.split(':')[:2])
        h2, m2 = map(int, t2.split(':')[:2])
        return abs((h2 * 60 + m2) - (h1 * 60 + m1))
    except (ValueError, IndexError):
        return None


def safe_float(val) -> Optional[float]:
    if val is None:
        return None
    try:
        s = str(val).replace(',', '').replace('(', '-').replace(')', '').strip()
        if not s or s == '-':
            return None
        return float(s)
    except (ValueError, TypeError):
        return None


def map_columns(headers: list) -> dict:
    mapping = {}
    for i, h in enumerate(headers):
        clean = str(h).strip()
        for key, pattern in COL_PATTERNS.items():
            if pattern.match(clean) and key not in mapping:
                mapping[key] = i
    return mapping


def parse_row(row: list, col_map: dict) -> Optional[dict]:
    def get(key):
        idx = col_map.get(key)
        if idx is not None and idx < len(row):
            return str(row[idx]).strip()
        return None

    def get_num(key):
        return safe_float(get(key))

    symbol = get('symbol')
    if not symbol:
        return None
    if re.match(r'^(total|grand|sub|net|sum)', symbol, re.I):
        return None

    result = {'symbol': symbol}

    # Time
    time_val = get('time')
    if time_val:
        m = re.search(r'(\d{1,2}:\d{2})', time_val)
        result['time'] = m.group(1) if m else time_val

    # Date
    result['date'] = get('date')

    # Side
    side = get('side')
    if side:
        result['side'] = 'BUY' if re.match(r'^(b|buy|long)', side, re.I) else 'SELL'

    # Quantities and prices
    result['qty'] = get_num('qty')
    result['price'] = get_num('price')
    result['amount'] = get_num('amount')
    result['pnl'] = get_num('pnl')
    result['buy_qty'] = get_num('buy_qty')
    result['sell_qty'] = get_num('sell_qty')
    result['buy_price'] = get_num('buy_price')
    result['sell_price'] = get_num('sell_price')

    # Infer side from buy/sell qty
    if not result.get('side'):
        if result.get('buy_qty') and result['buy_qty'] > 0:
            result['side'] = 'BUY'
        elif result.get('sell_qty') and result['sell_qty'] > 0:
            result['side'] = 'SELL'

    if not result.get('qty'):
        result['qty'] = result.get('buy_qty') or result.get('sell_qty')
    if not result.get('price'):
        result['price'] = result.get('buy_price') if result.get('side') == 'BUY' else result.get('sell_price')

    # Option components
    strike = get('strike')
    opt_type = get('opt_type')
    if strike or opt_type:
        result['symbol'] = f"{symbol} {strike or ''} {opt_type or ''}".strip()

    return result


def pair_trades(raw_trades: list) -> list:
    groups = {}
    for t in raw_trades:
        key = t['symbol'].lower().strip()
        if key not in groups:
            groups[key] = []
        groups[key].append(t)

    paired = []
    for trades in groups.values():
        buys = [t for t in trades if t.get('side') == 'BUY']
        sells = [t for t in trades if t.get('side') == 'SELL']

        max_pairs = min(len(buys), len(sells))
        for i in range(max_pairs):
            buy, sell = buys[i], sells[i]
            qty = min(buy.get('qty') or 1, sell.get('qty') or 1)
            entry = buy.get('price') or 0
            exit_p = sell.get('price') or 0
            pnl = buy.get('pnl') or sell.get('pnl') or round((exit_p - entry) * qty, 2)

            paired.append({
                'index': 0,
                'time': buy.get('time') or sell.get('time') or '09:15',
                'symbol': buy.get('symbol') or sell.get('symbol'),
                'side': 'BUY',
                'qty': qty,
                'entry': entry,
                'exit': exit_p,
                'pnl': pnl,
                'cum_pnl': 0,
                'session': classify_session(buy.get('time') or sell.get('time') or '09:15'),
                'time_gap_minutes': None,
                'tag': 'win' if pnl >= 0 else 'loss',
                'label': 'Winner' if pnl >= 0 else 'Loser',
            })

        # Unpaired trades with known P&L
        unpaired = buys[max_pairs:] + sells[max_pairs:]
        for t in unpaired:
            if t.get('pnl') is not None:
                paired.append({
                    'index': 0,
                    'time': t.get('time') or '09:15',
                    'symbol': t['symbol'],
                    'side': t.get('side') or 'BUY',
                    'qty': t.get('qty') or 1,
                    'entry': t.get('price') or 0,
                    'exit': 0,
                    'pnl': t['pnl'],
                    'cum_pnl': 0,
                    'session': classify_session(t.get('time') or '09:15'),
                    'time_gap_minutes': None,
                    'tag': 'win' if t['pnl'] >= 0 else 'loss',
                    'label': 'Winner' if t['pnl'] >= 0 else 'Loser',
                })

    # If no pairing, try raw trades with P&L
    if not paired and any(t.get('pnl') is not None for t in raw_trades):
        for t in raw_trades:
            paired.append({
                'index': 0,
                'time': t.get('time') or '09:15',
                'symbol': t['symbol'],
                'side': t.get('side') or 'BUY',
                'qty': t.get('qty') or 1,
                'entry': t.get('price') or 0,
                'exit': 0,
                'pnl': t.get('pnl') or 0,
                'cum_pnl': 0,
                'session': classify_session(t.get('time') or '09:15'),
                'time_gap_minutes': None,
                'tag': 'win' if (t.get('pnl') or 0) >= 0 else 'loss',
                'label': 'Winner' if (t.get('pnl') or 0) >= 0 else 'Loser',
            })

    # Sort by time
    paired.sort(key=lambda t: t['time'].replace(':', '').zfill(4))

    # Set indices, cum_pnl, time gaps
    cum_pnl = 0
    for i, t in enumerate(paired):
        t['index'] = i
        cum_pnl += t['pnl']
        t['cum_pnl'] = round(cum_pnl, 2)
        if i > 0:
            t['time_gap_minutes'] = time_gap_minutes(paired[i - 1]['time'], t['time'])

    return paired


def calculate_kpis(trades: list) -> dict:
    if not trades:
        return dict(net_pnl=0, total_trades=0, wins=0, losses=0, win_rate=0,
                     profit_factor=0, best_trade_pnl=0, worst_trade_pnl=0,
                     gross_profit=0, gross_loss=0, avg_win=0, avg_loss=0)

    wins = [t for t in trades if t['pnl'] > 0]
    losses = [t for t in trades if t['pnl'] <= 0]
    gross_profit = sum(t['pnl'] for t in wins)
    gross_loss = sum(t['pnl'] for t in losses)
    net_pnl = gross_profit + gross_loss

    return {
        'net_pnl': round(net_pnl, 2),
        'total_trades': len(trades),
        'wins': len(wins),
        'losses': len(losses),
        'win_rate': round(len(wins) / len(trades) * 100, 2) if trades else 0,
        'profit_factor': round(gross_profit / abs(gross_loss), 2) if gross_loss != 0 else (999 if wins else 0),
        'best_trade_pnl': max(t['pnl'] for t in trades),
        'worst_trade_pnl': min(t['pnl'] for t in trades),
        'gross_profit': round(gross_profit, 2),
        'gross_loss': round(gross_loss, 2),
        'avg_win': round(gross_profit / len(wins), 2) if wins else 0,
        'avg_loss': round(gross_loss / len(losses), 2) if losses else 0,
    }


def calculate_time_analysis(trades: list) -> dict:
    gaps = [t['time_gap_minutes'] for t in trades if t.get('time_gap_minutes') and t['time_gap_minutes'] > 0]
    times = []
    for t in trades:
        try:
            parts = t['time'].split(':')
            times.append(int(parts[0]) * 60 + int(parts[1]))
        except (ValueError, IndexError):
            pass

    return {
        'avg_time_gap_minutes': round(sum(gaps) / len(gaps), 1) if gaps else 0,
        'min_time_gap_minutes': min(gaps) if gaps else 0,
        'max_time_gap_minutes': max(gaps) if gaps else 0,
        'trading_duration_minutes': max(times) - min(times) if len(times) >= 2 else 0,
    }


# ═══════════════════════════════════════════
#   FILE TYPE PARSERS
# ═══════════════════════════════════════════

def parse_csv_file(filepath: str) -> tuple:
    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
        text = f.read()

    reader = csv.reader(io.StringIO(text))
    rows_list = list(reader)
    if len(rows_list) < 2:
        return text, []

    headers = rows_list[0]
    col_map = map_columns(headers)
    if len(col_map) < 2:
        return text, []

    trades = []
    for row in rows_list[1:]:
        trade = parse_row(row, col_map)
        if trade and trade.get('symbol'):
            trades.append(trade)

    return text, trades


def parse_excel_file(filepath: str) -> tuple:
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        return '', []

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    all_text = ''
    all_trades = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_list = []
        for row in ws.iter_rows(values_only=True):
            rows_list.append([str(c) if c is not None else '' for c in row])
            all_text += ' '.join(str(c) for c in row if c) + '\n'

        if len(rows_list) < 2:
            continue

        # Find header row
        header_idx = 0
        for i in range(min(10, len(rows_list))):
            non_empty = sum(1 for c in rows_list[i] if c.strip())
            if non_empty >= 3:
                header_idx = i
                break

        headers = rows_list[header_idx]
        col_map = map_columns(headers)
        if len(col_map) < 2:
            continue

        for row in rows_list[header_idx + 1:]:
            trade = parse_row(row, col_map)
            if trade and trade.get('symbol'):
                all_trades.append(trade)

    wb.close()
    return all_text, all_trades


def parse_pdf_file(filepath: str) -> tuple:
    try:
        import pdfplumber
    except ImportError:
        print("ERROR: pdfplumber not installed. Run: pip install pdfplumber", file=sys.stderr)
        return '', []

    all_text = ''
    all_trades = []

    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            # Extract text
            page_text = page.extract_text() or ''
            all_text += page_text + '\n'

            # Extract tables (pdfplumber is great at this)
            tables = page.extract_tables()
            for table in tables:
                if not table or len(table) < 2:
                    continue

                # First row as headers
                headers = [str(c or '').strip() for c in table[0]]
                col_map = map_columns(headers)

                if len(col_map) < 2:
                    # Try second row as header
                    if len(table) > 2:
                        headers = [str(c or '').strip() for c in table[1]]
                        col_map = map_columns(headers)
                        table = table[1:]

                if len(col_map) >= 2:
                    for row in table[1:]:
                        cells = [str(c or '').strip() for c in row]
                        trade = parse_row(cells, col_map)
                        if trade and trade.get('symbol'):
                            all_trades.append(trade)

    # If table extraction didn't work, try text parsing
    if not all_trades and all_text:
        lines = [l.strip() for l in all_text.split('\n') if l.strip()]

        # Try tab-delimited
        tab_lines = [l for l in lines if '\t' in l]
        if len(tab_lines) >= 3:
            reader = csv.reader(io.StringIO('\n'.join(tab_lines)), delimiter='\t')
            rows_list = list(reader)
            if rows_list:
                headers = rows_list[0]
                col_map = map_columns(headers)
                if len(col_map) >= 2:
                    for row in rows_list[1:]:
                        trade = parse_row(row, col_map)
                        if trade and trade.get('symbol'):
                            all_trades.append(trade)

        # Try multi-space delimited
        if not all_trades:
            spaced = [l for l in lines if re.search(r'\s{2,}', l) and len(re.split(r'\s{2,}', l)) >= 4]
            if len(spaced) >= 3:
                rows_list = [re.split(r'\s{2,}', l) for l in spaced]
                headers = rows_list[0]
                col_map = map_columns(headers)
                if len(col_map) >= 2:
                    for row in rows_list[1:]:
                        trade = parse_row(row, col_map)
                        if trade and trade.get('symbol'):
                            all_trades.append(trade)

    return all_text, all_trades


def parse_image_file(filepath: str) -> tuple:
    """Try OCR with pytesseract. Falls back to empty if not available."""
    try:
        import pytesseract
        from PIL import Image
        img = Image.open(filepath)
        text = pytesseract.image_to_string(img)
        if text.strip():
            # Try to parse the OCR text as CSV-like
            lines = [l.strip() for l in text.split('\n') if l.strip()]
            spaced = [l for l in lines if len(re.split(r'\s{2,}', l)) >= 3]
            if len(spaced) >= 3:
                rows_list = [re.split(r'\s{2,}', l) for l in spaced]
                headers = rows_list[0]
                col_map = map_columns(headers)
                if len(col_map) >= 2:
                    trades = []
                    for row in rows_list[1:]:
                        trade = parse_row(row, col_map)
                        if trade and trade.get('symbol'):
                            trades.append(trade)
                    return text, trades
        return text, []
    except ImportError:
        print("WARNING: pytesseract/Pillow not installed. Skipping image OCR.", file=sys.stderr)
        return '', []
    except Exception as e:
        print(f"WARNING: Image OCR failed: {e}", file=sys.stderr)
        return '', []


# ═══════════════════════════════════════════
#   MAIN ENTRY POINT
# ═══════════════════════════════════════════

def parse_trade_file(filepath: str) -> dict:
    """Parse a trade file and return structured JSON result."""
    ext = os.path.splitext(filepath)[1].lower().lstrip('.')
    raw_text = ''
    raw_trades = []

    try:
        if ext in ('csv', 'tsv'):
            raw_text, raw_trades = parse_csv_file(filepath)
        elif ext in ('xlsx', 'xls'):
            raw_text, raw_trades = parse_excel_file(filepath)
        elif ext == 'pdf':
            raw_text, raw_trades = parse_pdf_file(filepath)
        elif ext in ('png', 'jpg', 'jpeg', 'gif', 'webp'):
            raw_text, raw_trades = parse_image_file(filepath)
        else:
            # Try as text
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                raw_text = f.read()
            reader = csv.reader(io.StringIO(raw_text))
            rows_list = list(reader)
            if len(rows_list) >= 2:
                headers = rows_list[0]
                col_map = map_columns(headers)
                if len(col_map) >= 2:
                    for row in rows_list[1:]:
                        trade = parse_row(row, col_map)
                        if trade and trade.get('symbol'):
                            raw_trades.append(trade)

    except Exception as e:
        return {
            'success': False,
            'broker': detect_broker(raw_text),
            'market': detect_market(raw_text),
            'trade_date': detect_date(raw_text),
            'currency': detect_currency(raw_text),
            'total_trades_in_file': 0,
            'kpis': calculate_kpis([]),
            'trades': [],
            'time_analysis': calculate_time_analysis([]),
            'error': str(e),
        }

    if not raw_trades:
        return {
            'success': False,
            'broker': detect_broker(raw_text),
            'market': detect_market(raw_text),
            'trade_date': detect_date(raw_text),
            'currency': detect_currency(raw_text),
            'total_trades_in_file': 0,
            'kpis': calculate_kpis([]),
            'trades': [],
            'time_analysis': calculate_time_analysis([]),
            'error': f'Could not extract structured trades from {ext.upper()} file',
        }

    paired = pair_trades(raw_trades)
    kpis = calculate_kpis(paired)
    time_analysis = calculate_time_analysis(paired)

    return {
        'success': True,
        'broker': detect_broker(raw_text),
        'market': detect_market(raw_text),
        'trade_date': detect_date(raw_text),
        'currency': detect_currency(raw_text),
        'total_trades_in_file': len(paired),
        'kpis': kpis,
        'trades': paired,
        'time_analysis': time_analysis,
    }


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python parse_trades.py <file_path> [--output result.json]")
        sys.exit(1)

    filepath = sys.argv[1]
    output_file = None
    if '--output' in sys.argv:
        idx = sys.argv.index('--output')
        if idx + 1 < len(sys.argv):
            output_file = sys.argv[idx + 1]

    if not os.path.exists(filepath):
        print(f"ERROR: File not found: {filepath}", file=sys.stderr)
        sys.exit(1)

    print(f"Parsing: {filepath}")
    result = parse_trade_file(filepath)

    if result['success']:
        print(f"✅ {result['total_trades_in_file']} trades extracted from {result['broker']}")
        print(f"   Net P&L: {result['kpis']['net_pnl']} {result['currency']}")
        print(f"   Win Rate: {result['kpis']['win_rate']}%")
        print(f"   Profit Factor: {result['kpis']['profit_factor']}")
    else:
        print(f"⚠️  Parsing failed: {result.get('error', 'Unknown error')}")
        print("   (AI fallback should be used for this file)")

    json_output = json.dumps(result, indent=2, ensure_ascii=False)

    if output_file:
        with open(output_file, 'w') as f:
            f.write(json_output)
        print(f"\nSaved to: {output_file}")
    else:
        print(f"\n{json_output}")
